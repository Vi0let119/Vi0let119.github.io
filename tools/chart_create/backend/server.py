"""
Flask 后端 — 图表生成工具
AI 代理 + 图表 HTML 生成
"""
import os
import json
import re
import secrets
from io import BytesIO
from typing import Optional

from dotenv import load_dotenv
load_dotenv()

from flask import Flask, request, jsonify, session, send_file
from openai import OpenAI
from openpyxl import load_workbook

from skills.data_parser import parse_text_data, parse_excel_data
from skills.chart_recommender import recommend_chart_type
from skills.color_themes import MINIMAL_PALETTE, get_palette
from skills.html_builder import build_chart_html

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)

# ── CORS ──────────────────────────────────
@app.after_request
def add_cors_headers(resp):
    resp.headers['Access-Control-Allow-Origin'] = '*'
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return resp

@app.before_request
def handle_options():
    if request.method == 'OPTIONS':
        return '', 204

# ═══════════════════════════════════════════
# AI Provider 配置
# ═══════════════════════════════════════════

PROVIDERS = {
    'deepseek': {
        'name': 'DeepSeek',
        'base_url': 'https://api.deepseek.com/v1',
        'default_model': 'deepseek-chat',
        'api_key_env': 'DEEPSEEK_API_KEY',
        'type': 'openai',  # OpenAI 兼容
    },
    'qwen': {
        'name': '通义千问',
        'base_url': 'https://dashscope.aliyuncs.com/compatible-mode/v1',
        'default_model': 'qwen-plus',
        'api_key_env': 'QWEN_API_KEY',
        'type': 'openai',
    },
    'openai': {
        'name': 'OpenAI',
        'base_url': 'https://api.openai.com/v1',
        'default_model': 'gpt-4o-mini',
        'api_key_env': 'OPENAI_API_KEY',
        'type': 'openai',
    },
    'gemini': {
        'name': 'Gemini',
        'base_url': 'https://generativelanguage.googleapis.com/v1beta',
        'default_model': 'gemini-1.5-flash',
        'api_key_env': 'GEMINI_API_KEY',
        'type': 'gemini',
    },
}

# 从环境变量读取站点默认 API key
SITE_API_KEYS = {}
for key, cfg in PROVIDERS.items():
    env_val = os.getenv(cfg['api_key_env'], '')
    if env_val:
        SITE_API_KEYS[key] = env_val

# ═══════════════════════════════════════════
# AI System Prompt
# ═══════════════════════════════════════════

SYSTEM_PROMPT = """你是一个图表生成专家助手。你的任务是根据用户提供的数据和需求，生成高质量的 ECharts 图表。

## 工作流程

### 第一步：理解数据
- 分析用户发送的数据：列数、每列类型（数值/分类/时间）、数据范围、缺失情况
- 如果数据不明确，**先追问**再继续。例如：
  - "检测到数据有3列，请问哪一列作为X轴？"
  - "第2列是数值数据，我推断它作为Y轴数据，是否正确？"

### 第二步：推荐图表类型
根据数据特征推荐最合适的图表：
- 分类对比 → 柱状图 (bar)
- 时间/序列趋势 → 折线图 (line)
- 占比/比例 → 饼图 (pie)
- 两个数值变量关系 → 散点图 (scatter)

### 第三步：生成图表
输出完整的 ECharts option 配置，包裹在 ```echarts``` 代码块中。

## 设计规范（简约风）
- 使用柔和的低饱和度配色
- 背景干净，留白充足
- 字体清晰，标注简洁
- 配色参考：["#8aa4b0", "#b8a9a0", "#9bb5a0", "#c4a882", "#8b9a9e", "#b0a098"]

## 输出格式
分析完成后，必须输出：
1. 简要的分析结论
2. 图表推荐的依据
3. ```echarts
   {完整的 ECharts option JSON}
   ```

如果数据不足以生成图表，只输出追问问题，不要编造数据。"""

# ═══════════════════════════════════════════
# 辅助函数
# ═══════════════════════════════════════════

def get_client(provider: str, api_key: str = None):
    """获取 AI 客户端"""
    cfg = PROVIDERS.get(provider)
    if not cfg:
        raise ValueError(f"不支持的 AI 平台: {provider}")

    key = api_key or SITE_API_KEYS.get(provider)
    if not key:
        raise ValueError(f"未配置 {cfg['name']} 的 API Key")

    if cfg['type'] == 'openai':
        return OpenAI(api_key=key, base_url=cfg['base_url']), cfg['default_model']
    elif cfg['type'] == 'gemini':
        return _GeminiClient(key), cfg['default_model']
    raise ValueError(f"未知的 provider 类型: {cfg['type']}")


class _GeminiClient:
    """Gemini API 简易封装（兼容 OpenAI chat.completions 接口）"""
    def __init__(self, api_key):
        self.api_key = api_key
        self.base = 'https://generativelanguage.googleapis.com/v1beta'

    @property
    def chat(self):
        return self

    @property
    def completions(self):
        return self

    def create(self, model, messages, **kwargs):
        import requests
        # 转换 OpenAI 消息格式到 Gemini 格式
        contents = []
        system_text = ""
        for msg in messages:
            if msg['role'] == 'system':
                system_text = msg['content']
            else:
                role = 'user' if msg['role'] == 'user' else 'model'
                contents.append({'role': role, 'parts': [{'text': msg['content']}]})

        body = {
            'contents': contents,
            'generationConfig': {
                'temperature': kwargs.get('temperature', 0.7),
                'maxOutputTokens': kwargs.get('max_tokens', 4096),
            }
        }
        if system_text:
            body['systemInstruction'] = {'parts': [{'text': system_text}]}

        resp = requests.post(
            f"{self.base}/models/{model}:generateContent?key={self.api_key}",
            json=body
        )
        resp.raise_for_status()
        data = resp.json()

        class Response:
            class Choice:
                class Message:
                    def __init__(self, content):
                        self.content = content
                def __init__(self, msg):
                    self.message = self.Message(msg)
            def __init__(self, text):
                self.choices = [self.Choice(text)]

        text = data['candidates'][0]['content']['parts'][0]['text']
        return Response(text)


def extract_echarts_option(text: str) -> Optional[dict]:
    """从 AI 回复中提取 ECharts option"""
    # 匹配 ```echarts ... ``` 代码块
    pattern = r'```echarts\s*\n(.*?)\n\s*```'
    match = re.search(pattern, text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1))
        except json.JSONDecodeError:
            pass

    # 尝试匹配 ```json ... ```
    pattern = r'```json\s*\n(.*?)\n\s*```'
    match = re.search(pattern, text, re.DOTALL)
    if match:
        try:
            obj = json.loads(match.group(1))
            # 判断是不是 ECharts option（有 series 字段）
            if 'series' in obj:
                return obj
        except json.JSONDecodeError:
            pass

    return None


# ═══════════════════════════════════════════
# API 路由
# ═══════════════════════════════════════════

@app.route('/api/providers', methods=['GET'])
def list_providers():
    """列出可用 AI 平台"""
    result = []
    for key, cfg in PROVIDERS.items():
        result.append({
            'key': key,
            'name': cfg['name'],
            'has_site_key': key in SITE_API_KEYS,
        })
    return jsonify({'providers': result})


@app.route('/api/chat', methods=['POST'])
def chat():
    """AI 聊天接口"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请求体不能为空'}), 400

    message = data.get('message', '').strip()
    provider = data.get('provider', 'deepseek')
    user_api_key = data.get('api_key', '')
    conversation_history = data.get('history', [])
    file_content = data.get('file_content', '')

    if not message and not file_content:
        return jsonify({'error': '消息不能为空'}), 400

    # 构建消息
    user_message = message
    if file_content:
        data_section = f"\n\n[用户上传的数据]\n{file_content}"
        user_message = message + data_section if message else f"请分析以下数据：{data_section}"

    messages = [{'role': 'system', 'content': SYSTEM_PROMPT}]
    for h in conversation_history[-20:]:  # 保留最近 20 条
        messages.append({'role': h.get('role', 'user'), 'content': h.get('content', '')})
    messages.append({'role': 'user', 'content': user_message})

    try:
        client, model = get_client(provider, user_api_key or None)
        response = client.chat.completions.create(
            model=model,
            messages=messages,
            temperature=0.7,
            max_tokens=4096,
        )
        reply = response.choices[0].message.content

        # 提取 ECharts option
        option = extract_echarts_option(reply)

        return jsonify({
            'reply': reply,
            'option': option,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate-html', methods=['POST'])
def generate_html():
    """根据 ECharts option 生成完整 HTML 文件"""
    data = request.get_json()
    if not data or 'option' not in data:
        return jsonify({'error': '缺少 option 参数'}), 400

    option = data['option']
    title = data.get('title', '图表')
    html = build_chart_html(option, title)

    return jsonify({'html': html})


@app.route('/api/download-html', methods=['POST'])
def download_html():
    """生成并下载 HTML 文件"""
    data = request.get_json()
    if not data or 'option' not in data:
        return jsonify({'error': '缺少 option 参数'}), 400

    option = data['option']
    title = data.get('title', '图表')

    html = build_chart_html(option, title)
    html_bytes = ('﻿' + html).encode('utf-8')

    safe_name = re.sub(r'[<>:"/\\|?*]', '_', title)
    return send_file(
        BytesIO(html_bytes),
        mimetype='text/html;charset=UTF-8',
        as_attachment=True,
        download_name=f'{safe_name}.html',
    )


@app.route('/api/parse-file', methods=['POST'])
def parse_file():
    """解析上传的文件（txt/csv/xlsx）"""
    if 'file' not in request.files:
        return jsonify({'error': '未上传文件'}), 400

    file = request.files['file']
    filename = file.filename.lower() if file.filename else ''

    try:
        if filename.endswith(('.xlsx', '.xls')):
            wb = load_workbook(file, read_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append('\t'.join(str(c) if c is not None else '' for c in row))
            content = '\n'.join(rows)
            parsed = parse_excel_data(rows)
        else:
            content = file.read().decode('utf-8')
            parsed = parse_text_data(content)

        return jsonify({
            'content': content,
            'parsed': parsed,
        })
    except Exception as e:
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 400


@app.route('/api/recommend-chart', methods=['POST'])
def recommend_chart():
    """根据数据推荐图表类型"""
    data = request.get_json()
    if not data or 'columns' not in data:
        return jsonify({'error': '缺少 columns 参数'}), 400

    recommendation = recommend_chart_type(data['columns'])
    return jsonify(recommendation)


# ═══════════════════════════════════════════
# 启动
# ═══════════════════════════════════════════

if __name__ == '__main__':
    print("图表生成后端启动中...")
    print(f"已配置站点 API: {list(SITE_API_KEYS.keys())}")
    print("访问 http://localhost:5000")
    app.run(debug=True, port=5000)
